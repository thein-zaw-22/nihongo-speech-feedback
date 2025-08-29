import os
import json
import boto3
import tempfile
from botocore.exceptions import ClientError
from openai import OpenAI
import google.generativeai as genai
import logging
from logging import Formatter
import time
from collections import deque, defaultdict
from django.shortcuts import render, redirect
from django.utils import timezone
from datetime import datetime as dt_datetime, time as dt_time
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import UserCreationForm
from django.contrib.auth import login as auth_login
from django.contrib import messages as dj_messages
from .forms import AudioUploadForm, ProfileForm, ProfileAvatarForm, PasswordUpdateForm
from django.http import HttpResponse, JsonResponse, Http404
from django.core.files.base import ContentFile
import threading
import io
import csv
from .models import Transcription
from .models import GrammarQuestion, GrammarChoice, GrammarGameSession
from .models import BatchJob

# Configure logger
logger = logging.getLogger(__name__)

# --- LLM Provider Configuration ---
LLM_PROVIDER = os.getenv("LLM_PROVIDER", "openai").lower()
logger.info("Using LLM provider: %s", LLM_PROVIDER)

# Model configurations
BEDROCK_MODEL_ID = os.getenv("BEDROCK_MODEL_ID", "us.amazon.nova-lite-v1:0")
GEMINI_MODEL_ID = os.getenv("GEMINI_MODEL_ID", "gemini-2.5-flash-lite")

# Configure Gemini once
if os.getenv('GEMINI_API_KEY'):
    genai.configure(api_key=os.getenv('GEMINI_API_KEY'))

# ---------------- Rate limiting & retry (batch) ----------------
# Simple sliding window limiter per provider, configurable via env vars
class _RateLimiter:
    def __init__(self, max_calls: int, per_seconds: float):
        self.max_calls = max_calls
        self.per_seconds = per_seconds
        self._hits = deque()

    def wait(self):
        now = time.monotonic()
        # drop old
        while self._hits and (now - self._hits[0]) > self.per_seconds:
            self._hits.popleft()
        if len(self._hits) >= self.max_calls:
            sleep_for = self.per_seconds - (now - self._hits[0]) + 0.005
            if sleep_for > 0:
                time.sleep(sleep_for)
        # record
        self._hits.append(time.monotonic())

def _env_int(name, default):
    try:
        return int(os.getenv(name, str(default)))
    except Exception:
        return default

def _env_float(name, default):
    try:
        return float(os.getenv(name, str(default)))
    except Exception:
        return default

_limiters = {
    'openai': _RateLimiter(
        max_calls=_env_int('OPENAI_MAX_CALLS', 2),
        per_seconds=_env_float('OPENAI_PER_SECONDS', 1.0)
    ),
    'gemini': _RateLimiter(
        max_calls=_env_int('GEMINI_MAX_CALLS', 1),
        per_seconds=_env_float('GEMINI_PER_SECONDS', 2.0)
    ),
    'bedrock': _RateLimiter(
        max_calls=_env_int('BEDROCK_MAX_CALLS', 2),
        per_seconds=_env_float('BEDROCK_PER_SECONDS', 1.0)
    ),
}

def _call_llm_with_limits(text: str, provider: str):
    provider = (provider or LLM_PROVIDER).lower()
    limiter = _limiters.get(provider)
    if limiter:
        limiter.wait()
    # basic retry with backoff for transient 429/throughput issues
    attempts = _env_int('LLM_MAX_RETRIES', 3)
    base_delay = _env_float('LLM_RETRY_BASE_DELAY', 1.5)
    for i in range(attempts):
        try:
            return get_llm_feedback(text, provider)
        except Exception as e:
            msg = str(e).lower()
            transient = any(term in msg for term in [
                '429', 'rate limit', 'quota', 'throttle', 'throughput', 'too many requests', 'provisioned'
            ])
            if i < attempts - 1 and transient:
                sleep_for = base_delay * (2 ** i)
                time.sleep(sleep_for)
                continue
            raise
# ----------------------------------------------------------------

# ---------------- RPM limits loaded from .env ----------------
# These are per-model RPM (requests per minute) caps.
RPM_LIMITS = {
    "openai:gpt-4o": int(os.getenv("OPENAI_GPT4O_RPM", 1000)),
    "openai:gpt-4o-mini": int(os.getenv("OPENAI_GPT4O_MINI_RPM", 3500)),
    "gemini:gemini-2.5-flash-lite": int(os.getenv("GEMINI_25_FLASH_LITE_RPM", 15)),
    "aws:nova-lite": int(os.getenv("AWS_NOVA_LITE_RPM", 1000)),
}

class _RPMWindowLimiter:
    def __init__(self):
        self._hits = defaultdict(deque)  # key -> deque[timestamps]

    def wait(self, key: str):
        rpm = RPM_LIMITS.get(key)
        if not rpm or rpm <= 0:
            return
        window = 60.0
        dq = self._hits[key]
        now = time.monotonic()
        # prune
        while dq and (now - dq[0]) >= window:
            dq.popleft()
        if len(dq) >= rpm:
            sleep_for = window - (now - dq[0]) + 0.005
            if sleep_for > 0:
                time.sleep(sleep_for)
            # prune again after sleep
            now = time.monotonic()
            while dq and (now - dq[0]) >= window:
                dq.popleft()
        dq.append(time.monotonic())

_rpm_limiter = _RPMWindowLimiter()

def _rpm_key(provider: str, model_id: str) -> str:
    p = (provider or '').lower()
    m = (model_id or '').lower()
    if p == 'bedrock' or p == 'aws':
        # Map any nova-lite variants to aws:nova-lite
        if 'nova-lite' in m:
            return 'aws:nova-lite'
        return f'aws:{m}'
    elif p == 'gemini':
        return f'gemini:{m}'
    elif p == 'openai':
        return f'openai:{m}'
    return f'{p}:{m}'

def get_client(provider):
    """Get client for the specified provider."""
    if provider == "openai":
        return OpenAI(api_key=os.getenv('OPENAI_API_KEY'))
    elif provider == "gemini":
        return genai.GenerativeModel(GEMINI_MODEL_ID)
    elif provider == "bedrock":
        session = boto3.session.Session()
        aws_region = (
            os.getenv("AWS_REGION_NAME")
            or os.getenv("AWS_REGION")
            or os.getenv("AWS_DEFAULT_REGION")
            or session.region_name
            or "us-east-1"
        )
        _client_kwargs = {"region_name": aws_region}
        if os.getenv("AWS_ACCESS_KEY_ID") and os.getenv("AWS_SECRET_ACCESS_KEY"):
            _client_kwargs.update({
                "aws_access_key_id": os.getenv("AWS_ACCESS_KEY_ID"),
                "aws_secret_access_key": os.getenv("AWS_SECRET_ACCESS_KEY"),
            })
            if os.getenv("AWS_SESSION_TOKEN"):
                _client_kwargs["aws_session_token"] = os.getenv("AWS_SESSION_TOKEN")
        return boto3.client("bedrock-runtime", **_client_kwargs)
    else:
        raise ValueError(f"Unsupported LLM provider: {provider}")
# --------------------------------

# Whisper transcription support
try:
    import whisper
except Exception:
    whisper = None

_whisper_model = None

def get_whisper_model():
    global _whisper_model
    if whisper is None:
        raise RuntimeError("Whisper is not available in this environment")
    if _whisper_model is None:
        _whisper_model = whisper.load_model(os.getenv("WHISPER_MODEL", "base"))
    return _whisper_model

def transcribe_audio(uploaded_file):
    suffix = os.path.splitext(uploaded_file.name)[1] or ".wav"
    with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
        for chunk in uploaded_file.chunks():
            tmp.write(chunk)
        tmp_path = tmp.name
    try:
        model = get_whisper_model()
        result = model.transcribe(tmp_path, fp16=False, language='ja')
        return (result.get('text') or '').strip()
    finally:
        try:
            os.remove(tmp_path)
        except OSError:
            pass

def get_openai_feedback(transcript_text):
    """Gets feedback from OpenAI's GPT model."""
    client = get_client("openai")
    model_name = "gpt-4o-mini"
    _rpm_limiter.wait(_rpm_key('openai', model_name))
    messages = [
        {"role": "system", "content": (
            "You are a professional Japanese language tutor. "
            "Always correct unnatural or grammatical but awkward phrasing to the most natural Japanese usage. "
            "When I send a Japanese sentence, respond with only a JSON object with two keys: `corrected_text` and `corrections`. "
            "`corrected_text` must be the full corrected sentence. "
            "`corrections` must be an array of objects with `original`, `corrected`, and `explanation` fields. "
            "If the sentence is already the most natural usage, return it unchanged in `corrected_text` and an empty `corrections` array. "
            "Do not include extra text, markdown, or formatting. explanation should be English and concise. "
            "Example: {\"corrected_text\": \"水を飲みました。\", \"corrections\": "
            "[{\"original\": \"水は飲みました。\", \"corrected\": \"水を飲みました。\", "
            "\"explanation\": \"Use 'を' for the direct object instead of topic particle 'は'\"}]}"
        )},
        {"role": "user", "content": transcript_text}
    ]
    logger.debug("Sending messages to OpenAI: %s", messages)
    response = client.chat.completions.create(
        model=model_name,
        messages=messages
    )
    return response.choices[0].message.content

def get_gemini_feedback(transcript_text):
    """Gets feedback from Google Gemini model."""
    client = get_client("gemini")
    _rpm_limiter.wait(_rpm_key('gemini', GEMINI_MODEL_ID))
    prompt = (
        "You are a professional Japanese language tutor. "
        "Your task is to correct Japanese sentences to be more natural. "
        "Respond with ONLY a valid JSON object containing `corrected_text` and `corrections` keys. "
        "`corrected_text` should be the full, corrected sentence. "
        "`corrections` should be an array of objects, each with `original`, `corrected`, and `explanation` fields. "
        "The explanation must be concise and in English. If no correction is needed, return the original text and an empty array. "
        "Do not add any text before or after the JSON object.\n\n"
        f"Japanese text: {transcript_text}"
    )
    
    logger.debug("Sending prompt to Gemini (%s): %s", GEMINI_MODEL_ID, prompt[:100])
    try:
        response = client.generate_content(prompt)
        return response.text
    except Exception as e:
        logger.error("Gemini API error: %s", str(e))
        fallback = {
            "corrected_text": transcript_text,
            "corrections": [],
            "error": f"Gemini API error: {str(e)}"
        }
        return json.dumps(fallback)

def get_bedrock_feedback(transcript_text):
    """Gets feedback from Amazon Bedrock using the Converse API."""
    client = get_client("bedrock")
    _rpm_limiter.wait(_rpm_key('bedrock', BEDROCK_MODEL_ID))
    system_prompt = (
        "You are a professional Japanese language tutor. "
        "Your task is to correct Japanese sentences to be more natural. "
        "Respond with ONLY a valid JSON object containing `corrected_text` and `corrections` keys. "
        "`corrected_text` should be the full, corrected sentence. "
        "`corrections` should be an array of objects, each with `original`, `corrected`, and `explanation` fields. "
        "The explanation must be concise and in English. If no correction is needed, return the original text and an empty array. "
        "Do not add any text before or after the JSON object."
    )
    messages = [{"role": "user", "content": [{"text": transcript_text}]}]

    logger.debug(
        "Sending messages to Bedrock (%s, region=%s): %s",
        BEDROCK_MODEL_ID,
        os.getenv("AWS_REGION_NAME", "us-east-1"),
        messages,
    )
    try:
        response = client.converse(
            modelId=BEDROCK_MODEL_ID,
            messages=messages,
            system=[{"text": system_prompt}],
            inferenceConfig={"maxTokens": 300, "topP": 0.1, "temperature": 0.3},
            additionalModelRequestFields={"inferenceConfig": {"topK": 20}}
        )
    except ClientError as e:
        code = e.response.get("Error", {}).get("Code")
        if code == "AccessDeniedException":
            logger.error(
                "Access denied invoking Bedrock modelId=%s in region=%s. "
                "Ensure your account has model access enabled in Bedrock Console > Model access, "
                "and your IAM principal has bedrock:InvokeModel permissions.",
                BEDROCK_MODEL_ID,
                os.getenv("AWS_REGION_NAME", "us-east-1"),
            )
            # Return a structured feedback payload with error info instead of 500
            fallback = {
                "corrected_text": transcript_text,
                "corrections": [],
                "error": (
                    "Access denied to Bedrock model. Check Bedrock model access and IAM permissions."
                ),
            }
            return json.dumps(fallback)
        elif code == "ValidationException":
            message = e.response.get("Error", {}).get("Message", "")
            if "on-demand throughput" in message or "inference profile" in message:
                logger.error(
                    "Model %s requires an inference profile. Set BEDROCK_INFERENCE_PROFILE_ARN to a valid profile ARN or ID.",
                    BEDROCK_MODEL_ID,
                )
            elif "not authorized" in message:
                logger.error(
                    "Account not authorized for Bedrock API. Enable Bedrock access in AWS Console."
                )
            else:
                logger.error("Bedrock validation error: %s", message)
            fallback = {
                "corrected_text": transcript_text,
                "corrections": [],
                "error": (
                    "Bedrock access issue. Check account permissions and model access."
                ),
            }
            return json.dumps(fallback)
        # For other errors, re-raise for visibility
        raise

    raw_response = response['output']['message']['content'][0]['text']
    logger.debug("Raw feedback from Bedrock: %s", raw_response)
    return raw_response

def get_llm_feedback(transcript_text, provider=None):
    """Dispatches to the specified LLM provider."""
    provider = provider or LLM_PROVIDER
    if provider == "openai":
        return get_openai_feedback(transcript_text)
    elif provider == "gemini":
        return get_gemini_feedback(transcript_text)
    elif provider == "bedrock":
        return get_bedrock_feedback(transcript_text)
    # This should not be reached due to the initial check
    raise ValueError(f"Unsupported LLM provider: {provider}")

@login_required
def home(request):
    # Dashboard with cards to routes
    # Keep remembering provider selection for AI feedback card deep-link
    selected_provider = request.session.get('llm_provider')
    return render(request, 'home.html', {'selected_provider': selected_provider})


@login_required
def index(request):
    # Capture debug logs for this request
    logs = []
    class ListHandler(logging.Handler):
        def emit(self, record):
            logs.append(self.format(record))
    handler = ListHandler()
    handler.setLevel(logging.DEBUG)
    handler.setFormatter(Formatter('%(asctime)s - %(levelname)s - %(message)s'))
    logger.addHandler(handler)

    try:
        logger.debug("Index view called, method=%s", request.method)
        if request.method == 'POST':
            logger.debug("Received POST request for analysis")
            form = AudioUploadForm(request.POST, request.FILES)
            if form.is_valid():
                audio = form.cleaned_data.get('audio_file')
                text_input = (form.cleaned_data.get('text_input') or '').strip()
                if not text_input and audio:
                    try:
                        text_input = transcribe_audio(audio)
                        logger.debug("Transcribed audio to: %s", text_input[:100].replace('\n',' '))
                    except Exception as e:
                        logger.error("Transcription failed: %s", str(e))
                        return render(request, 'index.html', {'form': form, 'error': f'Transcription failed: {str(e)}'})

                if not text_input:
                    return render(request, 'index.html', {'form': form, 'error': 'Please enter text or upload audio for analysis.'})

                sentence = text_input
                selected_provider = form.cleaned_data.get('llm_provider', 'gemini')
                # Persist selection in session for future visits
                try:
                    request.session['llm_provider'] = selected_provider
                except Exception:
                    pass
                raw = get_llm_feedback(sentence, selected_provider)

                # Parse JSON feedback (with fallback slice)
                try:
                    feedback_data = json.loads(raw)
                except json.JSONDecodeError:
                    try:
                        start = raw.find('{')
                        end = raw.rfind('}') + 1
                        snippet = raw[start:end]
                        feedback_data = json.loads(snippet)
                    except Exception as e:
                        logger.warning("Fallback JSON parse failed: %s", str(e))
                        feedback_data = {"corrected_text": sentence, "corrections": [], "raw": raw}

                # Record which provider produced this feedback
                if isinstance(feedback_data, dict):
                    feedback_data['provider'] = selected_provider

                record = Transcription.objects.create(
                    user=request.user,
                    audio_file=audio if audio else None,
                    transcript=sentence,
                    feedback=feedback_data,
                )
                logger.debug("Saved Transcription id=%s", record.id)

                # Redirect to feedback page, carrying provider for display
                try:
                    # Persist a trimmed set of recent logs for display on feedback page
                    trimmed = logs[-100:]
                    request.session['last_logs'] = trimmed
                except Exception:
                    # Session may be unavailable or oversized; ignore quietly
                    pass
                return redirect(f"/feedback/{record.id}/?provider={selected_provider}")
        else:
            # Get provider from URL parameter if available
            selected_provider = request.GET.get('provider') or request.session.get('llm_provider', 'gemini')
            form = AudioUploadForm(initial={'llm_provider': selected_provider})

        return render(request, 'index.html', {'form': form, 'selected_provider': selected_provider})

    finally:
        logger.removeHandler(handler)


@login_required
def feedback(request, pk):
    try:
        obj = Transcription.objects.get(pk=pk)
    except Transcription.DoesNotExist:
        return redirect('index')
    selected_provider = request.GET.get('provider') or (obj.feedback.get('provider') if isinstance(obj.feedback, dict) else None)
    # Retrieve any carried-over debug logs from session (then clear)
    logs = None
    try:
        logs = request.session.pop('last_logs', None)
    except Exception:
        logs = None
    return render(request, 'result.html', {'result': obj, 'selected_provider': selected_provider, 'logs': logs})


@login_required
def history(request):
    items = Transcription.objects.filter(user=request.user).order_by('-created_at')
    start_str = request.GET.get('start')
    end_str = request.GET.get('end')
    tz = timezone.get_current_timezone()
    start_dt = end_dt = None
    # Parse YYYY-MM-DD safely
    try:
        if start_str:
            d = dt_datetime.strptime(start_str, '%Y-%m-%d').date()
            start_dt = timezone.make_aware(dt_datetime.combine(d, dt_time.min), tz)
    except Exception:
        start_str = None
    try:
        if end_str:
            d = dt_datetime.strptime(end_str, '%Y-%m-%d').date()
            end_dt = timezone.make_aware(dt_datetime.combine(d, dt_time.max), tz)
    except Exception:
        end_str = None
    # Validate range first: end must not be earlier than start
    if start_dt and end_dt and end_dt < start_dt:
        messages.error(request, 'End date cannot be earlier than start date.')
        # Show no results to make the error obvious
        items = Transcription.objects.none()
    else:
        # Apply filters only if valid
        if start_dt:
            items = items.filter(created_at__gte=start_dt)
        if end_dt:
            items = items.filter(created_at__lte=end_dt)
    return render(request, 'history.html', {'items': items, 'start': start_str, 'end': end_str})


@login_required
def flashcard(request):
    # Placeholder page for Flashcard feature
    return render(request, 'flashcard.html')


@login_required
def pronunciation(request):
    # Placeholder page for Pronunciation Coach
    return render(request, 'pronunciation.html')


@login_required
def grammar_game(request):
    # Placeholder page for Grammar Game
    return render(request, 'grammar_game.html')


@login_required
def grammar_play(request):
    # Returns JSON set of questions for selected level/category
    level = request.GET.get('level', 'N5')
    category = request.GET.get('category', 'particle')
    try:
        n = max(1, min(int(request.GET.get('n', '10')), 30))
    except Exception:
        n = 10
    qs = GrammarQuestion.objects.filter(jlpt_level=level, category=category, is_active=True).prefetch_related('choices').order_by('?')[:n]
    items = []
    for q in qs:
        choices = list(q.choices.all().values('text', 'is_correct'))
        items.append({
            'id': q.id,
            'prompt': q.prompt,
            'explanation': q.explanation,
            'jlpt_level': q.jlpt_level,
            'category': q.category,
            'choices': choices,
        })
    from django.http import JsonResponse
    return JsonResponse({'items': items})


@login_required
def grammar_submit(request):
    # Save session results (AJAX)
    from django.http import JsonResponse
    try:
        payload = json.loads(request.body.decode('utf-8'))
        level = payload.get('level', 'N5')
        category = payload.get('category', 'particle')
        total = int(payload.get('total', 0))
        correct = int(payload.get('correct', 0))
        duration = int(payload.get('duration', 0))
        best_streak = int(payload.get('best_streak', 0))
        timer_enabled = bool(payload.get('timer_enabled', False))
        detail_data = {
            'best_streak': best_streak,
            'timer_enabled': timer_enabled,
        }
        GrammarGameSession.objects.create(
            user=request.user,
            jlpt_level=level,
            category=category,
            total_questions=total,
            correct=correct,
            duration_seconds=duration,
            details=detail_data
        )
        return JsonResponse({'ok': True})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
def grammar_explain(request):
    from django.http import JsonResponse
    try:
        payload = json.loads(request.body.decode('utf-8'))
        prompt = payload.get('prompt', '')
        level = payload.get('level', '')
        category = payload.get('category', '')
        user_answer = payload.get('user_answer', '')
        correct_text = payload.get('correct_text', '')
        # Build a dedicated prompt asking for rich JSON
        system = (
            "You are a concise Japanese grammar coach. "
            "Return ONLY a JSON object with keys: summary (1-2 sentences), why_correct, why_incorrect, "
            "grammar_point (name + 1 line), examples (array of 2 short natural examples), tips (array of 2-3 short tips). "
            "Use clear English; include Japanese in examples. Do not include any extra text outside JSON."
        )
        user = (
            f"JLPT level: {level}\nCategory: {category}\nQuestion: {prompt}\n"
            f"User answer: {user_answer}\nCorrect answer: {correct_text}"
        )
        provider = request.session.get('llm_provider') or LLM_PROVIDER
        try:
            if provider == 'openai':
                client = get_client('openai')
                _rpm_limiter.wait(_rpm_key('openai', 'gpt-4o-mini'))
                resp = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[{"role":"system","content":system},{"role":"user","content":user}]
                )
                raw = resp.choices[0].message.content
            elif provider == 'gemini':
                client = get_client('gemini')
                _rpm_limiter.wait(_rpm_key('gemini', GEMINI_MODEL_ID))
                raw = client.generate_content(system + "\n\n" + user).text
            elif provider == 'bedrock':
                client = get_client('bedrock')
                _rpm_limiter.wait(_rpm_key('bedrock', BEDROCK_MODEL_ID))
                resp = client.converse(
                    modelId=BEDROCK_MODEL_ID,
                    messages=[{"role":"user","content":[{"text":user}]}],
                    system=[{"text":system}],
                    inferenceConfig={"maxTokens":400, "temperature":0.3}
                )
                raw = resp['output']['message']['content'][0]['text']
            else:
                raw = json.dumps({"summary":"Unsupported provider","why_correct":"","why_incorrect":"","grammar_point":"","examples":[],"tips":[]})
            # Parse JSON (attempt to extract braces if provider adds text)
            try:
                obj = json.loads(raw)
            except Exception:
                start = raw.find('{'); end = raw.rfind('}')+1
                obj = json.loads(raw[start:end]) if start >=0 and end>start else {"summary": raw}
        except Exception as e:
            obj = {"summary": f"Explanation unavailable: {e}", "why_correct":"", "why_incorrect":"", "grammar_point":"", "examples":[], "tips":[]}
        return JsonResponse({'ok': True, 'explanation': obj})
    except Exception as e:
        return JsonResponse({'ok': False, 'error': str(e)}, status=400)


@login_required
def grammar_history(request):
    # Filter per-user sessions
    qs = GrammarGameSession.objects.filter(user=request.user).order_by('-created_at')
    level = request.GET.get('level')
    category = request.GET.get('category')
    if level:
        qs = qs.filter(jlpt_level=level)
    if category:
        qs = qs.filter(category=category)
    best = qs.order_by('-details__best_streak', '-correct').first()
    return render(request, 'grammar_history.html', {
        'items': qs,
        'filter_level': level or '',
        'filter_category': category or '',
        'best': best,
        'levels': ['N5','N4','N3','N2','N1'],
    })


def signup(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            auth_login(request, user)
            try:
                from .models import Profile as UserProfile
                UserProfile.objects.get_or_create(user=user)
            except Exception:
                pass
            return redirect('index')
    else:
        form = UserCreationForm()
    return render(request, 'registration/signup.html', {'form': form})


# ---------------- Batch processing (queued with background thread) ----------------
def _process_batch_job(job_id: int):
    try:
        job = BatchJob.objects.get(pk=job_id)
    except BatchJob.DoesNotExist:
        return
    job.status = 'running'
    job.processed_rows = 0
    job.save(update_fields=['status', 'processed_rows', 'updated_at'])

    name = (job.input_file.name or '').lower()
    provider = job.provider
    try:
        if name.endswith('.csv'):
            with job.input_file.open('rb') as f:
                data = f.read().decode('utf-8', errors='replace')
            rows = list(csv.reader(io.StringIO(data)))
            job.total_rows = len(rows)
            job.save(update_fields=['total_rows', 'updated_at'])
            out = io.StringIO()
            writer = csv.writer(out)
            for i, row in enumerate(rows):
                # Check cancel
                job.refresh_from_db(fields=['cancel_requested'])
                if job.cancel_requested:
                    job.status = 'canceled'
                    job.save(update_fields=['status', 'updated_at'])
                    return
                if len(row) < 9:
                    row = row + [''] * (9 - len(row))
                text = (row[5] or '').strip()
                # Always treat first row as header and set output headers
                if i == 0:
                    row[6] = 'CorrectedText'
                    row[7] = 'Explanation'
                    row[8] = 'Notes'
                elif text:
                    try:
                        raw = _call_llm_with_limits(text, provider)
                        try:
                            parsed = json.loads(raw)
                        except Exception:
                            start = raw.find('{'); end = raw.rfind('}') + 1
                            parsed = json.loads(raw[start:end]) if start >= 0 and end > start else {"corrected_text": text, "corrections": []}
                        row[6] = parsed.get('corrected_text') or ''
                        row[7] = json.dumps(parsed, ensure_ascii=False)
                        # If the model signaled an error in payload, surface it in Notes (I)
                        note = ''
                        if isinstance(parsed, dict) and parsed.get('error'):
                            note = str(parsed.get('error'))
                        row[8] = note
                    except Exception as e:
                        row[6] = text
                        err_text = str(e)
                        row[7] = json.dumps({"error": err_text, "raw": raw if 'raw' in locals() else ''}, ensure_ascii=False)
                        row[8] = err_text
                writer.writerow(row)
                job.processed_rows = i + 1
                job.save(update_fields=['processed_rows', 'updated_at'])
            content = out.getvalue().encode('utf-8')
            job.output_file.save(os.path.basename(job.input_file.name).rsplit('.', 1)[0] + '_corrected.csv', ContentFile(content), save=False)
        else:
            from openpyxl import load_workbook
            with job.input_file.open('rb') as f:
                with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(job.input_file.name)[1]) as tmp:
                    tmp.write(f.read())
                    tmp_path = tmp.name
            try:
                wb = load_workbook(tmp_path)
                ws = wb.active
                max_row = ws.max_row
                job.total_rows = max_row
                job.save(update_fields=['total_rows', 'updated_at'])
                for r in range(1, max_row + 1):
                    # Check cancel
                    job.refresh_from_db(fields=['cancel_requested'])
                    if job.cancel_requested:
                        job.status = 'canceled'
                        job.save(update_fields=['status', 'updated_at'])
                        return
                    val = ws.cell(row=r, column=6).value
                    text = (val or '') if val is not None else ''
                    header = (r == 1)
                    if header:
                        ws.cell(row=r, column=7, value='CorrectedText')
                        ws.cell(row=r, column=8, value='Explanation')
                        ws.cell(row=r, column=9, value='Notes')
                    else:
                        text = str(text).strip()
                        if text:
                            try:
                                raw = _call_llm_with_limits(text, provider)
                                try:
                                    parsed = json.loads(raw)
                                except Exception:
                                    start = raw.find('{'); end = raw.rfind('}') + 1
                                    parsed = json.loads(raw[start:end]) if start >= 0 and end > start else {"corrected_text": text, "corrections": []}
                                ws.cell(row=r, column=7, value=parsed.get('corrected_text') or '')
                                ws.cell(row=r, column=8, value=json.dumps(parsed, ensure_ascii=False))
                                if isinstance(parsed, dict) and parsed.get('error'):
                                    ws.cell(row=r, column=9, value=str(parsed.get('error')))
                            except Exception as e:
                                ws.cell(row=r, column=7, value=text)
                                err_text = str(e)
                                ws.cell(row=r, column=8, value=json.dumps({"error": err_text, "raw": raw if 'raw' in locals() else ''}, ensure_ascii=False))
                                ws.cell(row=r, column=9, value=err_text)
                    job.processed_rows = r
                    job.save(update_fields=['processed_rows', 'updated_at'])
                bio = io.BytesIO()
                wb.save(bio)
                bio.seek(0)
                job.output_file.save(os.path.basename(job.input_file.name).rsplit('.', 1)[0] + '_corrected.xlsx', ContentFile(bio.read()), save=False)
            finally:
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass
        job.status = 'done'
        job.save()
    except Exception as e:
        job.status = 'error'
        job.error_message = str(e)
        job.save(update_fields=['status', 'error_message', 'updated_at'])


@login_required
def batch_create(request):
    if request.method != 'POST':
        raise Http404
    upload = request.FILES.get('batch_file')
    provider = request.POST.get('provider') or request.session.get('llm_provider') or LLM_PROVIDER
    if not upload:
        # If normal form post, bounce back with message
        if 'application/json' in (request.headers.get('Accept') or '') or request.headers.get('x-requested-with') == 'XMLHttpRequest':
            return JsonResponse({'ok': False, 'error': 'No file uploaded'}, status=400)
        messages.error(request, 'Please upload a CSV or Excel file.')
        return redirect('index')
    job = BatchJob.objects.create(user=request.user, provider=provider, input_file=upload, status='pending')
    threading.Thread(target=_process_batch_job, args=(job.id,), daemon=True).start()
    # Return JSON for AJAX; redirect for normal form submission
    if 'application/json' in (request.headers.get('Accept') or '') or request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'ok': True, 'job_id': job.id})
    from django.urls import reverse
    try:
        messages.success(request, f'Batch job #{job.id} started.')
    except Exception:
        pass
    return redirect(f"{reverse('index')}?job_id={job.id}")


@login_required
def batch_status(request, pk: int):
    try:
        job = BatchJob.objects.get(pk=pk, user=request.user)
    except BatchJob.DoesNotExist:
        raise Http404
    return JsonResponse({
        'ok': True,
        'status': job.status,
        'processed': job.processed_rows,
        'total': job.total_rows,
        'percent': job.progress_percent,
        'error': job.error_message,
        'cancelable': job.status in ('pending','running') and not job.cancel_requested,
        'download_url': (None if job.status != 'done' else f"/speak-ai/batch/{job.id}/download"),
    })


@login_required
def batch_download(request, pk: int):
    try:
        job = BatchJob.objects.get(pk=pk, user=request.user)
    except BatchJob.DoesNotExist:
        raise Http404
    if job.status != 'done' or not job.output_file:
        raise Http404
    ext = os.path.splitext(job.output_file.name)[1].lower()
    if ext == '.csv':
        ctype = 'text/csv; charset=utf-8'
    else:
        ctype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    with job.output_file.open('rb') as fh:
        data = fh.read()
    resp = HttpResponse(data, content_type=ctype)
    resp['Content-Disposition'] = f'attachment; filename="{os.path.basename(job.output_file.name)}"'
    return resp


@login_required
def batch_cancel(request, pk: int):
    if request.method != 'POST':
        raise Http404
    try:
        job = BatchJob.objects.get(pk=pk, user=request.user)
    except BatchJob.DoesNotExist:
        raise Http404
    if job.status in ('pending','running') and not job.cancel_requested:
        job.cancel_requested = True
        job.save(update_fields=['cancel_requested', 'updated_at'])
    # Return JSON for AJAX; redirect back to history for normal form posts
    if 'application/json' in (request.headers.get('Accept') or '') or request.headers.get('x-requested-with') == 'XMLHttpRequest':
        return JsonResponse({'ok': True})
    from django.urls import reverse
    try:
        messages.success(request, f'Cancel requested for job #{job.id}.')
    except Exception:
        pass
    return redirect(reverse('batch_history'))


@login_required
def batch_history(request):
    jobs = BatchJob.objects.filter(user=request.user).order_by('-created_at')[:50]
    return render(request, 'batch_history.html', {'jobs': jobs})


@login_required
def profile(request):
    # Ensure profile exists
    from .models import Profile as UserProfile
    profile_obj, _ = UserProfile.objects.get_or_create(user=request.user)
    if request.method == 'POST':
        # Special-case: avatar removal posts only the remove flag
        if request.POST.get('remove_avatar'):
            avatar_form = ProfileAvatarForm({'remove_avatar': True}, instance=profile_obj)
            if avatar_form.is_valid():
                avatar_form.save()
                try:
                    dj_messages.success(request, 'Avatar removed.')
                except Exception:
                    pass
                return redirect('profile')
            # If somehow invalid, fall through to normal rendering
            user_form = ProfileForm(instance=request.user)
        else:
            user_form = ProfileForm(request.POST, instance=request.user)
            avatar_form = ProfileAvatarForm(request.POST, request.FILES, instance=profile_obj)
            if user_form.is_valid() and avatar_form.is_valid():
                user_form.save()
                avatar_form.save()
                try:
                    dj_messages.success(request, 'Profile updated successfully.')
                except Exception:
                    pass
                return redirect('profile')
    else:
        user_form = ProfileForm(instance=request.user)
        avatar_form = ProfileAvatarForm(instance=profile_obj)
    return render(request, 'profile.html', {'form': user_form, 'avatar_form': avatar_form, 'profile': profile_obj})


@login_required
def password_update(request):
    from django.contrib.auth import update_session_auth_hash
    if request.method == 'POST':
        form = PasswordUpdateForm(request.user, request.POST)
        if form.is_valid():
            new_pwd = form.cleaned_data['new_password'].strip()
            user = request.user
            user.set_password(new_pwd)
            user.save()
            # Sanity check
            if not user.check_password(new_pwd):
                try:
                    dj_messages.error(request, 'Password update failed. Please try again.')
                except Exception:
                    pass
                return render(request, 'registration/password_update.html', {'form': form})
            update_session_auth_hash(request, user)
            try:
                dj_messages.success(request, 'Password updated successfully.')
            except Exception:
                pass
            return redirect('profile')
    else:
        form = PasswordUpdateForm(request.user)
    return render(request, 'registration/password_update.html', {'form': form})


@login_required
def batch_correct(request):
    """Upload CSV/XLSX, process column F via LLM, write G/H, return file."""
    if request.method != 'POST':
        return redirect('index')

    upload = request.FILES.get('batch_file')
    provider = request.POST.get('provider') or request.session.get('llm_provider') or LLM_PROVIDER
    if not upload:
        messages.error(request, 'Please upload a CSV or Excel file.')
        return redirect('index')

    name = (upload.name or '').lower()
    try:
        if name.endswith('.csv'):
            data = upload.read().decode('utf-8', errors='replace')
            src = io.StringIO(data)
            reader = csv.reader(src)
            out = io.StringIO()
            writer = csv.writer(out)
            for i, row in enumerate(reader):
                # Ensure at least 8 columns
                if len(row) < 8:
                    row = row + [''] * (8 - len(row))
                text = (row[5] or '').strip()
                # Header detection
                is_header = (i == 0 and text.lower() == 'incorrecttext')
                if not is_header and text:
                    try:
                        raw = get_llm_feedback(text, provider)
                        try:
                            parsed = json.loads(raw)
                        except Exception:
                            start = raw.find('{'); end = raw.rfind('}') + 1
                            parsed = json.loads(raw[start:end]) if start >= 0 and end > start else {"corrected_text": text, "corrections": []}
                        corrected = parsed.get('corrected_text') or ''
                        row[6] = corrected
                        row[7] = json.dumps(parsed, ensure_ascii=False)
                    except Exception as e:
                        row[6] = text
                        row[7] = json.dumps({"error": str(e), "raw": raw if 'raw' in locals() else ''}, ensure_ascii=False)
                else:
                    # Keep/ensure headers for outputs if header row
                    if is_header:
                        row[6] = 'CorrectedText'
                        row[7] = 'Explanation'
                writer.writerow(row)
            content = out.getvalue().encode('utf-8')
            resp = HttpResponse(content, content_type='text/csv; charset=utf-8')
            base = name.rsplit('.', 1)[0]
            resp['Content-Disposition'] = f'attachment; filename="{base}_corrected.csv"'
            return resp
        else:
            from openpyxl import load_workbook
            # Save to temp and load via openpyxl
            with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(upload.name)[1]) as tmp:
                for chunk in upload.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name
            try:
                wb = load_workbook(tmp_path)
                ws = wb.active
                max_row = ws.max_row
                for r in range(1, max_row + 1):
                    cell = ws.cell(row=r, column=6)  # F
                    text = (cell.value or '') if cell.value is not None else ''
                    header = isinstance(text, str) and text.strip().lower() == 'incorrecttext'
                    if r == 1 and header:
                        ws.cell(row=r, column=7, value='CorrectedText')
                        ws.cell(row=r, column=8, value='Explanation')
                        continue
                    text = str(text).strip()
                    if not text:
                        continue
                    try:
                        raw = get_llm_feedback(text, provider)
                        try:
                            parsed = json.loads(raw)
                        except Exception:
                            start = raw.find('{'); end = raw.rfind('}') + 1
                            parsed = json.loads(raw[start:end]) if start >= 0 and end > start else {"corrected_text": text, "corrections": []}
                        corrected = parsed.get('corrected_text') or ''
                        ws.cell(row=r, column=7, value=corrected)
                        ws.cell(row=r, column=8, value=json.dumps(parsed, ensure_ascii=False))
                    except Exception as e:
                        ws.cell(row=r, column=7, value=text)
                        ws.cell(row=r, column=8, value=json.dumps({"error": str(e), "raw": raw if 'raw' in locals() else ''}, ensure_ascii=False))

                # Write to bytes
                bio = io.BytesIO()
                wb.save(bio)
                bio.seek(0)
                resp = HttpResponse(bio.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                base = os.path.splitext(upload.name)[0]
                resp['Content-Disposition'] = f'attachment; filename="{base}_corrected.xlsx"'
                return resp
            finally:
                try:
                    os.remove(tmp_path)
                except Exception:
                    pass
    except Exception as e:
        messages.error(request, f'Batch processing failed: {e}')
        return redirect('index')
